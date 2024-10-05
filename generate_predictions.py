# -*- coding: utf-8 -*-
"""
Predictive Maintenance Using Neural Network Outputs
The neural network model generates predictions that identify whether 
manufacturing equipment requires calibration based on input sensor
data.
These predictions provide insights into equipment health, 
allowing maintenance to perform calibration and fault correction. 
The predictions are designed to optimize equipment uptime and reduce unplanned maintenance.
 
Created on Mon Sep 30 19:20:47 2024
@author: mcall
"""
# Phase 3: Prediction Generation

 # Filename: _generate_predictions.py Purpose: Use the trained model to make predictions on new data and save the results to an Excel file with formatting.

import pandas as pd
from tensorflow.keras.models import load_model
import joblib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

                # Load the trained model and scaler
model = load_model('C:/Users/mcall/trained_neural_network.h5')
scaler = joblib.load('C:/Users/mcall/scaler.pkl')

      # Load new data for predictions
data = pd.read_csv('C:/Users/mcall/synthetic_sensor_data_90_days.csv')

           # One-hot encode categorical columns (ensure consistency with training)
categorical_columns = ['sensor_type', 'equipment_state', 'usage_pattern']
for col in categorical_columns:
    if col in data.columns:
        data = pd.get_dummies(data, columns=[col], drop_first=True)

                 # Load the expected column names from the training phase
with open('C:/Users/mcall/training_columns.txt', 'r') as file:
    training_columns = [line.strip() for line in file.readlines()]

                # Align the DataFrame with the training columns, filling missing columns with zeros
data = data.reindex(columns=training_columns, fill_value=0)

                # Features for prediction (exclude non-feature columns)
features = data.drop(columns=['timestamp', 'sensor_id', 'calibration_needed'], errors='ignore')

                # Normalize the features using the saved scaler
features_scaled = scaler.transform(features)

# Generate predictions
predictions = model.predict(features_scaled)
data['calibration_prediction'] = (predictions > 0.5).astype(int)

 # Save predictions to an Excel file
output_file = 'C:/Users/mcall/predictions_output.xlsx'
data.to_excel(output_file, index=False, sheet_name='Predictions')

# Add conditional formatting for easier interpretation
wb = load_workbook(output_file)
ws = wb['Predictions']

# Apply red fill for rows where calibration is predicted
red_fill = PatternFill(start_color="FF1111", end_color="FF1111", fill_type="solid")
for row in range(2, ws.max_row + 1):
    if ws[f'N{row}'].value == 1:  # Adjust column index as needed
        for cell in ws[row]:
            cell.fill = red_fill

wb.save(output_file)
print(f"Predictions saved to: {output_file} with conditional formatting.")

"""
                           Predictions Summary: 

                          Data Input and Preprocessing:
        
        The model takes in various sensor readings, such as pressure
        deviation, temperature, humidity, vibration, and equipment state, 
        as input features. The input data is preprocessed through:
        One-Hot Encoding: Categorical variables like sensor_type, 
        equipment_state, and usage_pattern are transformed into a
        format suitable for the neural network by creating binary
        columns.Feature Scaling: Numerical features are standardized using 
        a scaler (e.g., StandardScaler) to normalize the range of
        values and improve model performance.
        The processed data is then aligned with the training data's
        expected structure to ensure consistency in the input features.

                             Generating Predictions:
        
        The neural network processes the input data to compute probabilities
        for each equipment instance. These probabilities represent the 
        likelihood of requiring calibration.
        The output layer of the neural network uses a sigmoid activation 
        function, which produces a value between 0 and 1. This value 
        reflects the confidence level of the model regarding whether
        calibration is needed.
        A threshold (typically 0.5) is applied to the probabilities to
        convert them into binary predictions:
            
            1: Calibration needed.
            0: No calibration needed.

                                 Output Storage:
        
        The predictions, along with the original sensor data, are saved
        in an Excel file (predictions_output.xlsx). This file provides 
        a comprehensive view of the equipment’s current state and the 
        model's assessment of whether maintenance is required.
        Conditional formatting is applied to the output file to
        highlight rows where calibration is predicted (1). This 
        visual indicator helps maintenance teams quickly identify
        which equipment requires attention.

                   Highlights of the Prediction Capabilities

    Fault Detection:
        The predictions are based on learned patterns from
    the training data, allowing the model to identify potential faults
    in equipment operation.
    
    Proactive Maintenance: 
        By predicting the need for calibration, the 
    model aids in scheduling proactive maintenance, reducing the risk of
    equipment failure and optimizing production schedules.
    
    Real-Time Analysis: 
        The predictions can be integrated into real-time 
    monitoring systems, providing continuous updates on equipment health.
    
    Interpretability: 
        The prediction output includes both the raw sensor 
    data and the model’s decision, making it easier for maintenance teams
    to understand the context behind each prediction.

                         Visualizing the Predictions

    The output file (predictions_output.xlsx) includes columns for sensor
    readings and a final column (calibration_prediction) that shows whether
    calibration is needed (1 for needed, 0 for not needed).
    The Excel file uses conditional formatting to highlight equipment that
    requires immediate attention, providing a user-friendly overview for 
    quick decision-making.

                             Example Use Cases

    Equipment Calibration Alerts: The predictions indicate which machines are
    most likely in need of calibration, allowing maintenance teams to 
    prioritize and address these machines promptly.
    Maintenance Scheduling: By tracking the predictions over time, 
    maintenance teams can optimize their schedules based on equipment
    health, reducing unexpected downtime and improving overall 
    efficiency.
    Anomaly Detection: The prediction system can act as an early 
    warning mechanism for detecting anomalies in sensor readings, 
    signaling potential faults before they lead to equipment failure.

                              Conclusion:

The prediction system powered by the neural network model serves as a 
proactive maintenance tool, analyzing sensor data to forecast equipment 
calibration needs. Its ability to provide real-time, interpretable 
outputs supports maintenance teams in making data-driven decisions, 
enhancing equipment reliability and operational efficiency.
"""